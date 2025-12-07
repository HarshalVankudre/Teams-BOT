"""
File Export Service for Teams Bot
Generates Excel and PDF files from equipment data.
"""
import io
import base64
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass

# Excel generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


@dataclass
class ExportResult:
    """Result of file export operation"""
    success: bool
    file_data: Optional[bytes] = None
    file_name: Optional[str] = None
    mime_type: Optional[str] = None
    base64_data: Optional[str] = None
    error: Optional[str] = None
    record_count: int = 0


class FileExportService:
    """Service for exporting data to Excel and PDF formats"""

    # Column display names (German)
    COLUMN_LABELS = {
        'id': 'ID',
        'bezeichnung': 'Bezeichnung',
        'hersteller': 'Hersteller',
        'kategorie': 'Kategorie',
        'geraetegruppe': 'Geraetegruppe',
        'seriennummer': 'Seriennummer',
        'inventarnummer': 'Inventarnummer',
        'kostenstelle': 'Kostenstelle',
        'verwendung': 'Verwendung',
        'gewicht_kg': 'Gewicht (kg)',
        'motor_leistung_kw': 'Leistung (kW)',
        'klimaanlage': 'Klimaanlage',
        'zentralschmierung': 'Zentralschmierung',
        'prop_arbeitsbreite': 'Arbeitsbreite',
        'prop_motor_hersteller': 'Motor-Hersteller',
        'prop_abgasstufe_eu': 'Abgasstufe EU',
    }

    # Priority columns to show first
    PRIORITY_COLUMNS = [
        'id', 'bezeichnung', 'hersteller', 'geraetegruppe',
        'seriennummer', 'kostenstelle', 'verwendung', 'gewicht_kg', 'motor_leistung_kw'
    ]

    def __init__(self):
        self.styles = getSampleStyleSheet()

    def _prepare_data(self, data: List[Dict[str, Any]]) -> Tuple[List[str], List[List[Any]]]:
        """Prepare data for export, ordering columns sensibly"""
        if not data:
            return [], []

        # Get all columns from data
        all_columns = set()
        for row in data:
            all_columns.update(row.keys())

        # Remove internal columns
        all_columns.discard('_query_context')
        all_columns.discard('_result_count')
        all_columns.discard('_results')

        # Order columns: priority first, then alphabetically
        ordered_columns = []
        for col in self.PRIORITY_COLUMNS:
            if col in all_columns:
                ordered_columns.append(col)
                all_columns.discard(col)

        # Add remaining columns sorted
        ordered_columns.extend(sorted(all_columns))

        # Build rows
        rows = []
        for item in data:
            row = []
            for col in ordered_columns:
                value = item.get(col)
                # Format special values
                if value is None:
                    value = ''
                elif isinstance(value, bool):
                    value = 'Ja' if value else 'Nein'
                elif isinstance(value, dict):
                    value = str(value.get('wert', ''))
                row.append(value)
            rows.append(row)

        return ordered_columns, rows

    def _get_display_headers(self, columns: List[str]) -> List[str]:
        """Convert column names to display labels"""
        return [self.COLUMN_LABELS.get(col, col.replace('prop_', '').replace('_', ' ').title())
                for col in columns]

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        title: str = "RUEKO Export",
        sheet_name: str = "Daten"
    ) -> ExportResult:
        """
        Export data to Excel file.

        Args:
            data: List of dictionaries to export
            title: Title for the export
            sheet_name: Name of the Excel sheet

        Returns:
            ExportResult with file data
        """
        try:
            if not data:
                return ExportResult(success=False, error="Keine Daten zum Exportieren")

            # Handle nested results format
            flat_data = self._flatten_results(data)

            columns, rows = self._prepare_data(flat_data)
            headers = self._get_display_headers(columns)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=f"{title} - {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            # Headers (row 3)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Data rows
            for row_idx, row in enumerate(rows, 4):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # Auto-adjust column widths (skip merged title row)
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                # Check header row (row 3) and data rows (row 4+)
                for row_idx in range(3, len(rows) + 4):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # Freeze header row
            ws.freeze_panes = 'A4'

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()

            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"rueko_export_{timestamp}.xlsx"

            return ExportResult(
                success=True,
                file_data=file_data,
                file_name=file_name,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                base64_data=base64.b64encode(file_data).decode('utf-8'),
                record_count=len(rows)
            )

        except Exception as e:
            return ExportResult(success=False, error=f"Excel-Export fehlgeschlagen: {str(e)}")

    def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        title: str = "RUEKO Export",
        max_columns: int = 8
    ) -> ExportResult:
        """
        Export data to PDF file.

        Args:
            data: List of dictionaries to export
            title: Title for the export
            max_columns: Maximum columns to include (PDF has space limits)

        Returns:
            ExportResult with file data
        """
        try:
            if not data:
                return ExportResult(success=False, error="Keine Daten zum Exportieren")

            # Handle nested results format
            flat_data = self._flatten_results(data)

            columns, rows = self._prepare_data(flat_data)

            # Limit columns for PDF (space constraints)
            if len(columns) > max_columns:
                columns = columns[:max_columns]
                rows = [row[:max_columns] for row in rows]

            headers = self._get_display_headers(columns)

            # Create PDF in landscape for more columns
            output = io.BytesIO()
            doc = SimpleDocTemplate(
                output,
                pagesize=landscape(A4),
                rightMargin=10*mm,
                leftMargin=10*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            elements = []

            # Title
            title_style = ParagraphStyle(
                'Title',
                parent=self.styles['Heading1'],
                fontSize=16,
                spaceAfter=10*mm,
                alignment=1  # Center
            )
            elements.append(Paragraph(f"{title}", title_style))

            # Subtitle with date and count
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=self.styles['Normal'],
                fontSize=10,
                spaceAfter=5*mm,
                alignment=1
            )
            elements.append(Paragraph(
                f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')} | {len(rows)} Datensaetze",
                subtitle_style
            ))
            elements.append(Spacer(1, 5*mm))

            # Build table data
            table_data = [headers] + rows

            # Calculate column widths based on content
            available_width = landscape(A4)[0] - 20*mm
            col_width = available_width / len(headers)

            # Create table
            table = Table(table_data, colWidths=[col_width] * len(headers))

            # Table style
            style = TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),

                # Data style
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),

                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),

                # Padding
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ])
            table.setStyle(style)

            elements.append(table)

            # Build PDF
            doc.build(elements)
            output.seek(0)
            file_data = output.read()

            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"rueko_export_{timestamp}.pdf"

            return ExportResult(
                success=True,
                file_data=file_data,
                file_name=file_name,
                mime_type="application/pdf",
                base64_data=base64.b64encode(file_data).decode('utf-8'),
                record_count=len(rows)
            )

        except Exception as e:
            return ExportResult(success=False, error=f"PDF-Export fehlgeschlagen: {str(e)}")

    def _flatten_results(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Flatten nested result format from agent system"""
        flat = []
        for item in data:
            if '_results' in item:
                # Nested format from SQL agent
                flat.extend(item['_results'])
            else:
                flat.append(item)
        return flat


# Global instance
file_export_service = FileExportService()
